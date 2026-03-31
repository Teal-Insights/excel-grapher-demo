"""
LibreOffice headless recalc sanity check for GDP forecast shocks.

Edits the workbook copy with ``gdp_forecast_value_from_bps``, converts via
``soffice --headless --convert-to xlsx``, reads Figure 1 Chart Data cells, and
optionally compares those values to a Python ``FormulaEvaluator`` payload
(``build_figure1_payload`` / precache JSON) for the same bps levels.

``top_deltas`` / ``max_abs_delta`` are **internal to LibreOffice** (shock minus
baseline under LO only). When ``python_baseline_payload`` and
``python_shock_payload`` are provided, ``python_vs_libreoffice`` reports whether
Python matches LO at each level and whether the **shock increment** matches:
``(py_shock - py_base) - (lo_shock - lo_base)`` per cell.

Note: LO edits every workbook-discovered GDP forecast cell; the graph evaluator
may only override a subset. Residual error can reflect that input mismatch as
well as formula-engine differences.
"""

from __future__ import annotations

import math
import os
import shutil
import subprocess
import tempfile
import json
from pathlib import Path
from typing import Any

import openpyxl

from .payload import (
    CHART_SHEET,
    FIGURE1_PANELS,
    _read_gdp_forecast_cell_values_from_workbook,
    col_letters,
    gdp_forecast_value_from_bps,
)


def parse_qualified_a1(qualified: str) -> tuple[str, str]:
    if "!" not in qualified:
        raise ValueError(f"Not a qualified A1 reference: {qualified!r}")
    sheet_part, addr = qualified.rsplit("!", 1)
    sheet_part = sheet_part.strip()
    addr = addr.strip()
    if sheet_part.startswith("'") and sheet_part.endswith("'"):
        sheet_part = sheet_part[1:-1].replace("''", "'")
    return sheet_part, addr


def find_soffice(explicit: str | None = None) -> str | None:
    if explicit:
        p = Path(explicit)
        if p.is_file():
            return str(p)
        w = shutil.which(explicit)
        return w
    for name in ("soffice", "libreoffice"):
        w = shutil.which(name)
        if w:
            return w
    return None


def load_gdp_input_targets(workbook: Path) -> list[tuple[str, str, float]]:
    got = _read_gdp_forecast_cell_values_from_workbook(workbook)
    if not got:
        raise RuntimeError(
            f"No GDP forecast cells found in {workbook} (see lic_dsf.payload GDP_FORECAST_*)."
        )
    keys, vals = got
    targets: list[tuple[str, str, float]] = []
    for k, v in zip(keys, vals, strict=True):
        sheet, a1 = parse_qualified_a1(k)
        base = float(v) if v is not None else 0.0
        targets.append((sheet, a1, base))
    return targets


def write_shocked_xlsm(
    src: Path,
    dst: Path,
    targets: list[tuple[str, str, float]],
    bps: int,
) -> None:
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dst)
    wb = openpyxl.load_workbook(dst, keep_vba=True)
    for sheet, a1, base in targets:
        wb[sheet][a1] = gdp_forecast_value_from_bps(base, bps)
    wb.save(dst)


def libreoffice_to_xlsx(
    xlsm: Path,
    outdir: Path,
    *,
    soffice: str,
    timeout_s: int,
) -> Path:
    outdir.mkdir(parents=True, exist_ok=True)
    xlsm = xlsm.resolve()
    subprocess.run(
        [
            soffice,
            "--headless",
            "--norestore",
            "--nodefault",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(outdir.resolve()),
            str(xlsm),
        ],
        check=True,
        timeout=timeout_s,
        stdout=subprocess.DEVNULL,
        env={**os.environ, "HOME": os.environ.get("HOME", str(Path.home()))},
    )
    return outdir / f"{xlsm.stem}.xlsx"


def _cell_numeric(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, bool):
        return float(v)
    if isinstance(v, (int, float)):
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
        return float(v)
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def read_figure1_chart_values(xlsx: Path) -> dict[str, float | None]:
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    try:
        ws = wb[CHART_SHEET]
    except KeyError as e:
        wb.close()
        raise RuntimeError(f"Sheet {CHART_SHEET!r} missing from {xlsx}") from e
    out: dict[str, float | None] = {}
    for panel in FIGURE1_PANELS:
        for spec in panel.series:
            for col in col_letters():
                a1 = f"{col}{spec.value_row}"
                q = f"{CHART_SHEET}!{a1}"
                out[q] = _cell_numeric(ws[a1].value)
    wb.close()
    return out


def payload_for_bps_from_precache_doc(doc: dict[str, Any], bps: int) -> dict[str, Any]:
    """Return Figure 1 ``payload`` from a ``figure1-gdp-shocks.json``-style document."""
    default = doc.get("default")
    if isinstance(default, dict) and int(default.get("bps", 0)) == bps:
        pl = default.get("payload")
        if isinstance(pl, dict):
            return pl
    for e in doc.get("shocks") or []:
        if not isinstance(e, dict):
            continue
        if int(e.get("bps", -999999)) == bps:
            pl = e.get("payload")
            if isinstance(pl, dict):
                return pl
    raise ValueError(f"No precache entry for bps={bps}")


def payloads_from_precache_json(
    path: Path,
    *,
    baseline_bps: int,
    shock_bps: int,
) -> tuple[dict[str, Any], dict[str, Any]]:
    doc = json.loads(path.read_text(encoding="utf-8"))
    return (
        payload_for_bps_from_precache_doc(doc, baseline_bps),
        payload_for_bps_from_precache_doc(doc, shock_bps),
    )


def figure1_payload_to_chart_map(payload: dict[str, Any]) -> dict[str, float | None]:
    """
    Flatten a Figure 1 payload to the same keys as ``read_figure1_chart_values``
    (``Chart Data!D61``, …).
    """
    panels = payload.get("panels") or []
    if len(panels) != len(FIGURE1_PANELS):
        raise ValueError(
            f"payload has {len(panels)} panels, expected {len(FIGURE1_PANELS)}"
        )
    out: dict[str, float | None] = {}
    for pi, panel_spec in enumerate(FIGURE1_PANELS):
        p = panels[pi]
        series_list = p.get("series") or []
        if len(series_list) != len(panel_spec.series):
            raise ValueError(
                f"panel {pi} has {len(series_list)} series, expected {len(panel_spec.series)}"
            )
        for si, spec in enumerate(panel_spec.series):
            s = series_list[si]
            data = list(s.get("data") or [])
            for ci, col in enumerate(col_letters()):
                addr = f"{CHART_SHEET}!{col}{spec.value_row}"
                raw = data[ci] if ci < len(data) else None
                out[addr] = _cell_numeric(raw)
    return out


def _compare_maps_python_minus_lo(
    python_map: dict[str, float | None],
    lo_map: dict[str, float | None],
    *,
    top_n: int,
) -> dict[str, Any]:
    rows: list[tuple[str, float, float, float]] = []
    for k in python_map:
        if k not in lo_map:
            continue
        pv, lv = python_map[k], lo_map[k]
        if pv is None or lv is None:
            continue
        rows.append((k, pv, lv, pv - lv))
    rows.sort(key=lambda r: abs(r[3]), reverse=True)
    errs = [r[3] for r in rows]
    max_abs = max(abs(e) for e in errs) if errs else None
    mean_abs = sum(abs(e) for e in errs) / len(errs) if errs else None
    top: list[dict[str, Any]] = []
    for k, pv, lv, diff in rows[:top_n]:
        top.append(
            {
                "cell": k,
                "python": pv,
                "libreoffice": lv,
                "python_minus_libreoffice": diff,
            }
        )
    return {
        "cells_compared": len(rows),
        "max_abs_error": max_abs,
        "mean_abs_error": mean_abs,
        "top_errors": top,
    }


def _compare_shock_increment_python_minus_lo(
    py_b: dict[str, float | None],
    py_s: dict[str, float | None],
    lo_b: dict[str, float | None],
    lo_s: dict[str, float | None],
    *,
    top_n: int,
) -> dict[str, Any]:
    rows: list[tuple[str, float, float, float]] = []
    for k in py_b:
        if k not in py_s or k not in lo_b or k not in lo_s:
            continue
        p0, p1, l0, l1 = py_b[k], py_s[k], lo_b[k], lo_s[k]
        if p0 is None or p1 is None or l0 is None or l1 is None:
            continue
        py_inc = p1 - p0
        lo_inc = l1 - l0
        rows.append((k, py_inc, lo_inc, py_inc - lo_inc))
    rows.sort(key=lambda r: abs(r[3]), reverse=True)
    errs = [r[3] for r in rows]
    max_abs = max(abs(e) for e in errs) if errs else None
    mean_abs = sum(abs(e) for e in errs) / len(errs) if errs else None
    top: list[dict[str, Any]] = []
    for k, pi, li, diff in rows[:top_n]:
        top.append(
            {
                "cell": k,
                "python_delta": pi,
                "libreoffice_delta": li,
                "python_delta_minus_lo_delta": diff,
            }
        )
    return {
        "cells_compared": len(rows),
        "max_abs_error": max_abs,
        "mean_abs_error": mean_abs,
        "top_errors": top,
    }


def diff_chart_maps(
    baseline: dict[str, float | None],
    shocked: dict[str, float | None],
) -> list[tuple[str, float | None, float | None, float | None]]:
    rows: list[tuple[str, float | None, float | None, float | None]] = []
    for k in baseline:
        vb = baseline[k]
        vs = shocked.get(k)
        if vb is None and vs is None:
            delta = None
        elif vb is None or vs is None:
            delta = None
        else:
            delta = vs - vb
        rows.append((k, vb, vs, delta))
    rows.sort(
        key=lambda r: (
            float("-inf") if r[3] is None else abs(r[3]),
            r[0],
        ),
        reverse=True,
    )
    return rows


def run_libreoffice_gdp_shock_check(
    workbook: Path,
    *,
    baseline_bps: int = 0,
    shock_bps: int = 10,
    timeout_s: int = 600,
    soffice: str | None = None,
    keep_temps: bool = False,
    top_n: int = 15,
    python_baseline_payload: dict[str, Any] | None = None,
    python_shock_payload: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """
    Run openpyxl edit + two LibreOffice exports and diff Chart Data Figure 1 cells.

    If ``python_baseline_payload`` and ``python_shock_payload`` are set (Figure 1
    ``{"categories","panels"}`` dicts from ``build_figure1_payload``), adds
    ``python_vs_libreoffice`` with baseline-level, shock-level, and incremental
    agreement stats (Python minus LibreOffice).

    Returns a JSON-serializable dict with keys: ok, baseline_bps, shock_bps,
    soffice, gdp_input_cells, output_cells_compared, max_abs_delta, mean_abs_delta,
    top_deltas (LO-only), python_vs_libreoffice (optional), temp_dir, error.
    """
    src = workbook.resolve()
    if not src.is_file():
        return {
            "ok": False,
            "error": f"Workbook not found: {src}",
        }

    bin_soffice = find_soffice(soffice)
    if not bin_soffice:
        return {
            "ok": False,
            "error": "LibreOffice not found (try --lo-soffice or PATH: soffice, libreoffice).",
        }

    tmp = Path(tempfile.mkdtemp(prefix="lo-gdp-shock-"))
    try:
        targets = load_gdp_input_targets(src)
        base_xlsm = tmp / f"{src.stem}_baseline_{baseline_bps}bps.xlsm"
        shock_xlsm = tmp / f"{src.stem}_shock_{shock_bps}bps.xlsm"

        write_shocked_xlsm(src, base_xlsm, targets, baseline_bps)
        write_shocked_xlsm(src, shock_xlsm, targets, shock_bps)

        base_xlsx = libreoffice_to_xlsx(
            base_xlsm, tmp, soffice=bin_soffice, timeout_s=timeout_s
        )
        shock_xlsx = libreoffice_to_xlsx(
            shock_xlsm, tmp, soffice=bin_soffice, timeout_s=timeout_s
        )

        if not base_xlsx.is_file() or not shock_xlsx.is_file():
            err: dict[str, Any] = {
                "ok": False,
                "error": (
                    f"Missing xlsx after convert: base={base_xlsx.is_file()}, "
                    f"shock={shock_xlsx.is_file()}"
                ),
            }
            if keep_temps:
                err["temp_dir"] = str(tmp)
            else:
                shutil.rmtree(tmp, ignore_errors=True)
            return err

        vb = read_figure1_chart_values(base_xlsx)
        vs = read_figure1_chart_values(shock_xlsx)
        rows = diff_chart_maps(vb, vs)

        finite_deltas = [r[3] for r in rows if r[3] is not None]
        max_abs = max(abs(d) for d in finite_deltas) if finite_deltas else None
        mean_abs = (
            sum(abs(d) for d in finite_deltas) / len(finite_deltas)
            if finite_deltas
            else None
        )

        top: list[dict[str, Any]] = []
        for addr, b0, b1, d in rows:
            if d is None:
                continue
            top.append(
                {
                    "cell": addr,
                    "baseline_value": b0,
                    "shock_value": b1,
                    "delta": d,
                }
            )
            if len(top) >= top_n:
                break

        out: dict[str, Any] = {
            "ok": True,
            "baseline_bps": baseline_bps,
            "shock_bps": shock_bps,
            "soffice": bin_soffice,
            "gdp_input_cells": len(targets),
            "output_cells_compared": len(rows),
            "libreoffice_internal": {
                "description": "LO recalc: shock minus baseline (same engine only)",
                "max_abs_delta": max_abs,
                "mean_abs_delta": mean_abs,
                "top_deltas": top,
            },
            "max_abs_delta": max_abs,
            "mean_abs_delta": mean_abs,
            "top_deltas": top,
        }

        if python_baseline_payload is not None and python_shock_payload is not None:
            try:
                py_b = figure1_payload_to_chart_map(python_baseline_payload)
                py_s = figure1_payload_to_chart_map(python_shock_payload)
                out["python_vs_libreoffice"] = {
                    "description": (
                        "FormulaEvaluator payload vs LibreOffice xlsx (same cell keys); "
                        "errors are python - libreoffice"
                    ),
                    f"at_{baseline_bps}_bps": _compare_maps_python_minus_lo(
                        py_b, vb, top_n=top_n
                    ),
                    f"at_{shock_bps}_bps": _compare_maps_python_minus_lo(
                        py_s, vs, top_n=top_n
                    ),
                    "shock_increment": _compare_shock_increment_python_minus_lo(
                        py_b, py_s, vb, vs, top_n=top_n
                    ),
                }
            except ValueError as e:
                out["python_vs_libreoffice"] = {
                    "error": f"Could not map Python payload: {e!s}",
                }
        if keep_temps:
            out["temp_dir"] = str(tmp)
        else:
            shutil.rmtree(tmp, ignore_errors=True)
        return out
    except (OSError, subprocess.CalledProcessError, RuntimeError) as e:
        if keep_temps:
            err_out: dict[str, Any] = {
                "ok": False,
                "error": repr(e),
                "temp_dir": str(tmp),
            }
        else:
            shutil.rmtree(tmp, ignore_errors=True)
            err_out = {"ok": False, "error": repr(e)}
        return err_out
    except BaseException:
        if not keep_temps:
            shutil.rmtree(tmp, ignore_errors=True)
        raise


def print_check_report(result: dict[str, Any]) -> None:
    """Human-readable lines for CLI tools."""
    if not result.get("ok"):
        print(f"LibreOffice check failed: {result.get('error', 'unknown')}")
        if result.get("temp_dir"):
            print(f"Temp dir: {result['temp_dir']}")
        return

    print(f"LibreOffice check: {result['soffice']}")
    print(f"  GDP forecast input cells: {result['gdp_input_cells']}")
    print(
        f"  LO internal: {result['baseline_bps']} bps vs {result['shock_bps']} bps "
        "(recalc delta under LibreOffice only)"
    )
    print(f"  Output cells compared: {result['output_cells_compared']}")
    ma = result.get("max_abs_delta")
    mn = result.get("mean_abs_delta")
    if ma is not None:
        print(f"  LO max |shock - baseline|: {ma:.6g}")
    if mn is not None:
        print(f"  LO mean |shock - baseline|: {mn:.6g}")
    print("  LO largest |shock - baseline| (sample):")
    for row in result.get("top_deltas") or []:
        print(
            f"    {row['cell']:28}  base={row['baseline_value']!s:>14}  "
            f"shock={row['shock_value']!s:>14}  Δ={row['delta']: .6g}"
        )

    pvl = result.get("python_vs_libreoffice")
    if pvl:
        if pvl.get("error"):
            print(f"  Python vs LibreOffice: {pvl['error']}")
        else:
            print("  Python (FormulaEvaluator) vs LibreOffice — same Chart Data cells:")
            for key in sorted(
                k for k in pvl if k not in ("description", "error")
            ):
                block = pvl[key]
                if not isinstance(block, dict):
                    continue
                if block.get("error"):
                    print(f"    {key}: {block['error']}")
                    continue
                mx = block.get("max_abs_error")
                mn_b = block.get("mean_abs_error")
                nc = block.get("cells_compared")
                if mx is not None and mn_b is not None:
                    print(
                        f"    {key}: cells={nc}  max|py-LO|={mx:.6g}  mean|py-LO|={mn_b:.6g}"
                    )
                else:
                    print(f"    {key}: cells={nc}")
                for row in (block.get("top_errors") or [])[:5]:
                    if "python_minus_libreoffice" in row:
                        print(
                            f"      {row['cell']}  py={row['python']:.6g}  "
                            f"LO={row['libreoffice']:.6g}  py-LO={row['python_minus_libreoffice']:.6g}"
                        )
                    elif "python_delta_minus_lo_delta" in row:
                        print(
                            f"      {row['cell']}  Δpy={row['python_delta']:.6g}  "
                            f"ΔLO={row['libreoffice_delta']:.6g}  "
                            f"Δpy-ΔLO={row['python_delta_minus_lo_delta']:.6g}"
                        )

    if result.get("temp_dir"):
        print(f"  Kept temp dir: {result['temp_dir']}")
